import pandas as pd
import numpy as np
import joblib
import os
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH  = os.path.join(BASE_DIR, "data", "nomination_synthetic_dataset_v3.xlsx")
MODELS_DIR = os.path.join(BASE_DIR, "models")
DB_PATH    = os.path.join(BASE_DIR, "nominations.db")

model    = joblib.load(os.path.join(MODELS_DIR, "best_model.pkl"))
features = joblib.load(os.path.join(MODELS_DIR, "feature_names.pkl"))

# ── Setup SQLite DB ───────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS programs (
            program_id    INTEGER PRIMARY KEY AUTOINCREMENT,
            program_name  TEXT,
            min_qual      TEXT,
            min_exp       INTEGER,
            training_week TEXT,
            batch_size    INTEGER,
            waitlist_size INTEGER,
            run_at        TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS nominations (
            nomination_id INTEGER PRIMARY KEY AUTOINCREMENT,
            program_id    INTEGER,
            employee_id   TEXT,
            employee_name TEXT,
            department    TEXT,
            qualification TEXT,
            experience    INTEGER,
            ai_score      REAL,
            fuzzy_match   REAL,
            status        TEXT,
            rank          INTEGER,
            FOREIGN KEY (program_id) REFERENCES programs(program_id)
        )
    """)
    conn.commit()
    conn.close()

def save_to_db(prog_meta, selected, waitlist):
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("""
        INSERT INTO programs
        (program_name, min_qual, min_exp, training_week, batch_size, waitlist_size, run_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        prog_meta["name"], prog_meta["min_qual_name"], prog_meta["min_exp"],
        prog_meta["week_name"], prog_meta["batch_size"], prog_meta["waitlist_n"],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    program_id = c.lastrowid
    for rank, (_, r) in enumerate(selected.iterrows(), start=1):
        c.execute("""
            INSERT INTO nominations
            (program_id, employee_id, employee_name, department, qualification,
             experience, ai_score, fuzzy_match, status, rank)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (program_id, r["Employee ID"], r["Name"], r["Department"],
              r["Qualification"], int(r["experience"]),
              float(r["ai_score"]), float(r["fuzzy_match"]), "Selected", rank))
    for rank, (_, r) in enumerate(waitlist.iterrows(), start=1):
        c.execute("""
            INSERT INTO nominations
            (program_id, employee_id, employee_name, department, qualification,
             experience, ai_score, fuzzy_match, status, rank)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (program_id, r["Employee ID"], r["Name"], r["Department"],
              r["Qualification"], int(r["experience"]),
              float(r["ai_score"]), float(r["fuzzy_match"]), "Waitlist", rank))
    conn.commit()
    conn.close()
    return program_id

def show_history():
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("SELECT * FROM programs ORDER BY run_at DESC")
    rows = c.fetchall()
    conn.close()
    if not rows:
        print("\n  No saved programs yet.")
        return
    print(f"\n  {'ID':<5} {'Program Name':<35} {'Week':<20} {'Batch':<7} {'Saved At'}")
    print("  " + "-" * 80)
    for row in rows:
        print(f"  {row[0]:<5} {row[1]:<35} {row[3]:<20} {row[5]:<7} {row[7]}")

# ── Export to Excel ───────────────────────────────────────────────────────────
def export_to_excel(prog_name, prog_meta, selected, waitlist, rejected):

    # helpers
    def hdr_style(cell, bg="1F4E79"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def body_style(cell, bold=False, color="000000", align="left"):
        cell.font = Font(bold=bold, color=color, name="Arial", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")

    def thin_border():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    # Title block
    ws.merge_cells("A1:B1")
    ws["A1"] = "NOMINATION FILTERING SYSTEM"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Shortlist Report — {prog_name}"
    ws["A2"].font = Font(bold=True, size=11, color="2E75B6", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.append([])  # blank row

    # Program details
    details = [
        ("Program Name",      prog_meta["name"]),
        ("Min Qualification", prog_meta["min_qual_name"]),
        ("Min Experience",    f"{prog_meta['min_exp']} years"),
        ("Training Week",     prog_meta["week_name"]),
        ("Batch Size",        prog_meta["batch_size"]),
        ("Waitlist Size",     prog_meta["waitlist_n"]),
        ("Generated On",      datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("Total Selected",    len(selected)),
        ("Total Waitlisted",  len(waitlist)),
        ("Total Rejected",    len(rejected)),
        ("Avg AI Score (Selected)",  f"{selected['ai_score'].mean():.2f}"),
        ("Avg Fuzzy Match (Selected)", f"{selected['fuzzy_match'].mean():.1f}%"),
    ]
    for label, value in details:
        ws.append([label, value])
        row = ws.max_row
        ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="404040")
        ws[f"B{row}"].font = Font(name="Arial", size=10)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="F2F2F2")
        ws.row_dimensions[row].height = 16

    # ── Sheet 2: Selected ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("✅ Selected")
    headers = ["Rank", "Employee ID", "Name", "Department", "Area of Work", "Qualification",
               "Experience (Yrs)", "AI Score", "Qual%", "Exp%",
               "Avail%", "Skill%", "Train%", "Area%", "Fuzzy Match%", "Fit"]
    widths  = [6, 13, 22, 14, 22, 14, 17, 11, 8, 8, 8, 8, 8, 8, 14, 14]

    ws2.append(headers)
    for i, (col, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=i)
        hdr_style(cell)
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 18
    ws2.freeze_panes = "A2"

    def fit_text(pct):
        if pct >= 75:   return "Strong Fit"
        elif pct >= 50: return "Moderate Fit"
        elif pct >= 30: return "Weak Fit"
        else:           return "Poor Fit"

    FIT_COLORS = {
        "Strong Fit":   ("14532D", "DCFCE7"),
        "Moderate Fit": ("713F12", "FEF9C3"),
        "Weak Fit":     ("7C2D12", "FFEDD5"),
        "Poor Fit":     ("7F1D1D", "FEE2E2"),
    }

    for rank, (_, r) in enumerate(selected.iterrows(), start=1):
        fit = fit_text(r["fuzzy_match"])
        row_data = [
            rank, r["Employee ID"], r["Name"], r["Department"],
            r["Area of Work"], r["Qualification"], int(r["experience"]),
            round(float(r["ai_score"]), 2),
            f"{r['fz_qual']:.0f}%", f"{r['fz_exp']:.0f}%",
            f"{r['fz_avail']:.0f}%", f"{r['fz_skill']:.0f}%",
            f"{r['fz_train']:.0f}%", f"{r['fz_area']:.0f}%",
            f"{r['fuzzy_match']:.1f}%", fit
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row
        bg = "FFFFFF" if rank % 2 == 0 else "F7FAFF"
        for col_idx, _ in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill    = PatternFill("solid", fgColor=bg)
            cell.font    = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left",
                                       vertical="center")
            cell.border  = thin_border()
        # Color the fit cell
        fit_cell = ws2.cell(row=row_idx, column=14)
        fc, bc   = FIT_COLORS.get(fit, ("000000", "FFFFFF"))
        fit_cell.fill = PatternFill("solid", fgColor=bc)
        fit_cell.font = Font(name="Arial", size=10, color=fc, bold=True)
        # Color AI score cell
        score_cell = ws2.cell(row=row_idx, column=7)
        score_cell.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        ws2.row_dimensions[row_idx].height = 16

    # ── Sheet 3: Waitlist ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("⏳ Waitlist")
    ws3.append(headers)
    for i, (col, w) in enumerate(zip(headers, widths), 1):
        cell = ws3.cell(row=1, column=i)
        hdr_style(cell, bg="7B3F00")
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[1].height = 18
    ws3.freeze_panes = "A2"

    for rank, (_, r) in enumerate(waitlist.iterrows(), start=1):
        fit = fit_text(r["fuzzy_match"])
        row_data = [
            rank, r["Employee ID"], r["Name"], r["Department"],
            r["Area of Work"], r["Qualification"], int(r["experience"]),
            round(float(r["ai_score"]), 2),
            f"{r['fz_qual']:.0f}%", f"{r['fz_exp']:.0f}%",
            f"{r['fz_avail']:.0f}%", f"{r['fz_skill']:.0f}%",
            f"{r['fz_train']:.0f}%", f"{r['fz_area']:.0f}%",
            f"{r['fuzzy_match']:.1f}%", fit
        ]
        ws3.append(row_data)
        row_idx = ws3.max_row
        bg = "FFFFFF" if rank % 2 == 0 else "FFFBF5"
        for col_idx, _ in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left",
                                       vertical="center")
            cell.border    = thin_border()
        ws3.row_dimensions[row_idx].height = 16

    # ── Save file ─────────────────────────────────────────────────────────────
    exports_dir = os.path.join(BASE_DIR, "exports")
    os.makedirs(exports_dir, exist_ok=True)
    safe_name   = "".join(c if c.isalnum() or c in " _-" else "_" for c in prog_name)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path   = os.path.join(exports_dir, f"Shortlist_{safe_name}_{timestamp}.xlsx")
    wb.save(file_path)
    return file_path

# ── Init DB ───────────────────────────────────────────────────────────────────
init_db()

print("=" * 65)
print("  NOMINATION FILTERING SYSTEM — SHORTLIST GENERATOR")
print("=" * 65)

show = input("\n  View past saved shortlists first? (y/n): ").strip().lower()
if show == "y":
    show_history()
    print()

# ── Load sheets ───────────────────────────────────────────────────────────────
df_emp    = pd.read_excel(DATA_PATH, sheet_name="Employees")
df_skills = pd.read_excel(DATA_PATH, sheet_name="Skills")
df_hist   = pd.read_excel(DATA_PATH, sheet_name="Training History")
df_avail  = pd.read_excel(DATA_PATH, sheet_name="Availability")

# ── User enters program details ───────────────────────────────────────────────
print("\n  Enter your training program details:")
print("  " + "-" * 45)
prog_name   = input("  Program name                        : ").strip()
print("  Min qualification: 1=High School  2=Diploma  3=Bachelor's  4=Master's  5=MBA  6=PhD")
min_qual    = int(input("  Enter number                        : ").strip())
min_exp     = int(input("  Minimum experience (years)          : ").strip())
print("  Training week: 1=Week1(Jun9-13)  2=Week2(Jun16-20)  3=Week3(Jun23-27)  4=Week4(Jun30-Jul4)")
week_choice = int(input("  Enter number                        : ").strip())
batch_size  = int(input("  Batch size                          : ").strip())
waitlist_n  = int(input("  Waitlist size (default 5)           : ").strip() or "5")

print("\n  Target areas of work (optional — press Enter to include ALL areas)")
print("  Example: Software Development, DevOps, Data Management")
area_input   = input("  Enter areas (comma separated)       : ").strip()
target_areas = [a.strip().lower() for a in area_input.split(",") if a.strip()] if area_input else []

QUAL_NAMES = {1:"High School",2:"Diploma",3:"Bachelor's",4:"Master's",5:"MBA",6:"PhD"}
WEEK_NAMES = {1:"Week 1 (Jun 9-13)",2:"Week 2 (Jun 16-20)",
              3:"Week 3 (Jun 23-27)",4:"Week 4 (Jun 30-Jul 4)"}

# ── Feature Engineering ───────────────────────────────────────────────────────
QUAL_RANK = {"High School":1,"Diploma":2,"Bachelor's":3,"Master's":4,"MBA":5,"PhD":6}
df_emp["qual_rank"]   = df_emp["Qualification"].map(QUAL_RANK).fillna(1)
df_emp["experience"]  = df_emp["Experience (Years)"]
df_emp["performance"] = df_emp["Performance Rating (1-5)"]
df_emp["Last Promotion Date"] = pd.to_datetime(
    df_emp["Last Promotion Date"], dayfirst=True, errors="coerce")
TODAY = pd.Timestamp("2025-06-05")
df_emp["days_since_promotion"] = (TODAY - df_emp["Last Promotion Date"]).dt.days.fillna(365)
dept_map = {d: i for i, d in enumerate(df_emp["Department"].unique())}
df_emp["dept_code"] = df_emp["Department"].map(dept_map).fillna(0)

hist_completed = df_hist[df_hist["Status"] == "Completed"]
hist_failed    = df_hist[df_hist["Status"].isin(["Failed","Dropped"])]
hist_agg = hist_completed.groupby("Employee ID").agg(
    trainings_completed=("Program Name","count"),
    avg_training_score=("Score (%)","mean")).reset_index()
hist_fail_agg = hist_failed.groupby("Employee ID").agg(
    trainings_failed=("Program Name","count")).reset_index()

LEVEL_RANK = {"Beginner":1,"Intermediate":2,"Advanced":3}
for col in ["Skill 1 Level","Skill 2 Level","Skill 3 Level"]:
    df_skills[col] = df_skills[col].map(LEVEL_RANK).fillna(1)
df_skills["avg_skill_level"] = df_skills[
    ["Skill 1 Level","Skill 2 Level","Skill 3 Level"]].mean(axis=1)
df_skills["max_skill_level"] = df_skills[
    ["Skill 1 Level","Skill 2 Level","Skill 3 Level"]].max(axis=1)
skill_agg = df_skills[["Employee ID","avg_skill_level","max_skill_level"]]

avail_cols  = [c for c in df_avail.columns if c != "Employee ID"]
week_col    = avail_cols[week_choice - 1]
for col in avail_cols:
    df_avail[col] = df_avail[col].map({"Yes":1,"No":0}).fillna(0)
df_avail["weeks_available"] = df_avail[avail_cols].sum(axis=1)
avail_agg = df_avail[["Employee ID","weeks_available", week_col]].copy()
avail_agg = avail_agg.rename(columns={week_col: "free_this_week"})

df = df_emp[["Employee ID","Name","Department","Area of Work","Qualification",
             "experience","qual_rank","performance",
             "days_since_promotion","dept_code"]].copy()
df = df.merge(hist_agg,      on="Employee ID", how="left")
df = df.merge(hist_fail_agg, on="Employee ID", how="left")
df = df.merge(skill_agg,     on="Employee ID", how="left")
df = df.merge(avail_agg,     on="Employee ID", how="left")
df = df.fillna(0)

# ── Stage 1: Hard Filters ─────────────────────────────────────────────────────
total       = len(df)
# Area of work filter
if target_areas:
    area_mask   = df["Area of Work"].str.lower().isin(target_areas)
    df_area     = df[area_mask].copy()
else:
    df_area     = df.copy()

df_eligible = df_area[
    (df_area["free_this_week"] == 1) &
    (df_area["qual_rank"]      >= min_qual) &
    (df_area["experience"]     >= min_exp)
].copy()

print(f"\n  ── Stage 1: Hard Filtering ──────────────────────")
print(f"  Total employees   : {total}")
if target_areas:
    print(f"  Area filter       : {', '.join(target_areas)}")
    print(f"  After area filter : {len(df_area)}")
print(f"  Filtered out      : {total - len(df_eligible)}")
print(f"  Eligible pool     : {len(df_eligible)}")

if len(df_eligible) == 0:
    print("\n  ⚠️  No eligible employees found.")
    exit()

if len(df_eligible) < batch_size:
    print(f"\n  ⚠️  Only {len(df_eligible)} eligible — selecting all.")
    batch_size = len(df_eligible)
    waitlist_n = 0

# ── Stage 2: AI Score ─────────────────────────────────────────────────────────
X = df_eligible[features]
df_eligible = df_eligible.copy()
df_eligible["ai_score"] = np.clip(model.predict(X), 10, 95).round(2)

# ── Stage 3: Fuzzy Match ──────────────────────────────────────────────────────
def fuzzy_qual(q, mq):
    gap = q - mq
    return 100.0 if gap >= 0 else (60.0 if gap==-1 else (30.0 if gap==-2 else 10.0))

def fuzzy_exp(e, me):
    if e >= me or me == 0: return 100.0
    return round(max(10.0, (e/me)**1.5*100), 1)

def fuzzy_avail(free, weeks):
    return 100.0 if free==1 else (30.0 if weeks>0 else 0.0)

def fuzzy_skill(avg, mx):
    return round((mx/3)*60 + (avg/3)*40, 1)

def fuzzy_train(tc, ts, tf):
    return round(max(0.0, min(tc/4,1.0)*50 + (ts/100)*35 - min(tf*10,30)), 1)

# Area of work fuzzy match
AREA_SKILL_MAP = {
    "software development" : ["Python","Java","SQL","Machine Learning","Cloud Computing","Agile"],
    "qa testing"           : ["Python","SQL","Agile","Communication"],
    "devops"               : ["Python","Cloud Computing","Agile","Cybersecurity"],
    "embedded systems"     : ["Python","Java","AutoCAD"],
    "data management"      : ["SQL","Power BI","Tableau","Data Analysis","Machine Learning"],
    "network administration": ["Cybersecurity","Cloud Computing"],
    "cybersecurity"        : ["Cybersecurity","Risk Management","Cloud Computing"],
    "it support"           : ["SQL","Communication","CRM Tools"],
    "digital marketing"    : ["Digital Marketing","SEO","Content Writing","Power BI"],
    "brand management"     : ["Digital Marketing","Content Writing","Communication","Public Speaking"],
    "content strategy"     : ["Content Writing","SEO","Communication","Public Speaking"],
    "market research"      : ["Data Analysis","Excel","Power BI","Tableau"],
    "accounting"           : ["Excel","Financial Modeling","SAP"],
    "financial analysis"   : ["Excel","Financial Modeling","Power BI","Tableau","SQL"],
    "budgeting"            : ["Excel","Financial Modeling","SAP"],
    "auditing"             : ["Excel","Risk Management","SAP"],
    "supply chain"         : ["Excel","Six Sigma","SAP","Project Management"],
    "logistics"            : ["Excel","Six Sigma","SAP"],
    "process improvement"  : ["Six Sigma","Project Management","Agile","Excel"],
    "procurement"          : ["Negotiation","Excel","SAP","CRM Tools"],
    "talent acquisition"   : ["Communication","CRM Tools","Excel","Negotiation"],
    "employee relations"   : ["Communication","Public Speaking","Leadership"],
    "training & development": ["Communication","Public Speaking","Leadership","Excel"],
    "compensation"         : ["Excel","Financial Modeling","SAP"],
    "b2b sales"            : ["Negotiation","CRM Tools","Communication","Public Speaking"],
    "account management"   : ["CRM Tools","Negotiation","Communication","Excel"],
    "pre-sales"            : ["Communication","Negotiation","Public Speaking","CRM Tools"],
    "channel sales"        : ["Negotiation","CRM Tools","Excel"],
    "contract management"  : ["Negotiation","Risk Management","Communication"],
    "compliance"           : ["Risk Management","Excel","Communication"],
    "corporate law"        : ["Negotiation","Communication","Risk Management"],
    "intellectual property": ["Negotiation","Communication"],
}

def fuzzy_area(area_of_work, prog_skill, target_areas):
    area_lower  = str(area_of_work).lower()
    prog_lower  = str(prog_skill).lower()
    related     = AREA_SKILL_MAP.get(area_lower, [])
    related_low = [s.lower() for s in related]
    # Exact skill match for this area
    if prog_lower in related_low:
        return 100.0
    # Partial — area is in target but skill not directly mapped
    if target_areas and area_lower in target_areas:
        return 70.0
    # Area not targeted but still in pool
    if not target_areas:
        return 50.0   # no filter applied — neutral
    return 20.0

df_eligible["fz_area"]  = df_eligible.apply(
    lambda r: fuzzy_area(r["Area of Work"], "general", target_areas), axis=1)

df_eligible["fz_qual"]  = df_eligible.apply(lambda r: fuzzy_qual(r["qual_rank"], min_qual), axis=1)
df_eligible["fz_exp"]   = df_eligible.apply(lambda r: fuzzy_exp(r["experience"], min_exp), axis=1)
df_eligible["fz_avail"] = df_eligible.apply(lambda r: fuzzy_avail(r["free_this_week"], r["weeks_available"]), axis=1)
df_eligible["fz_skill"] = df_eligible.apply(lambda r: fuzzy_skill(r["avg_skill_level"], r["max_skill_level"]), axis=1)
df_eligible["fz_train"] = df_eligible.apply(lambda r: fuzzy_train(
    r["trainings_completed"], r["avg_training_score"], r["trainings_failed"]), axis=1)
df_eligible["fuzzy_match"] = (
    df_eligible["fz_qual"]  * 0.18 +
    df_eligible["fz_exp"]   * 0.22 +
    df_eligible["fz_avail"] * 0.18 +
    df_eligible["fz_skill"] * 0.18 +
    df_eligible["fz_train"] * 0.12 +
    df_eligible["fz_area"]  * 0.12
).round(1)

# Sort by AI score (primary)
df_eligible = df_eligible.sort_values("ai_score", ascending=False).reset_index(drop=True)
selected    = df_eligible.iloc[:batch_size].copy()
waitlist    = df_eligible.iloc[batch_size:batch_size+waitlist_n].copy()
rejected    = df_eligible.iloc[batch_size+waitlist_n:].copy()

def fit_label(pct):
    if pct >= 75:   return "🟢 Strong"
    elif pct >= 50: return "🟡 Moderate"
    elif pct >= 30: return "🟠 Weak"
    else:           return "🔴 Poor"

def print_section(title, emoji, subset):
    print(f"\n{'='*90}")
    print(f"  {emoji}  {title}")
    print(f"{'='*90}")
    print(f"  {'#':<4} {'Name':<20} {'Area of Work':<22} {'AI Score':>9} "
          f"{'Qual%':>6} {'Exp%':>5} {'Avl%':>5} {'Skl%':>5} {'Trn%':>5} {'Area%':>6} "
          f"{'Match%':>7}  Fit")
    print(f"  {'-'*105}")
    for i, (_, r) in enumerate(subset.iterrows()):
        print(f"  {i+1:<4} {r['Name']:<20} {r['Area of Work']:<22} "
              f"{r['ai_score']:>8.2f}  "
              f"{r['fz_qual']:>5.0f}% {r['fz_exp']:>4.0f}% "
              f"{r['fz_avail']:>4.0f}% {r['fz_skill']:>4.0f}% "
              f"{r['fz_train']:>4.0f}% {r['fz_area']:>5.0f}%  "
              f"{r['fuzzy_match']:>5.1f}%  {fit_label(r['fuzzy_match'])}")

print_section(f"SELECTED ({len(selected)}) — {prog_name}", "✅", selected)
print_section(f"WAITLIST ({len(waitlist)})", "⏳", waitlist)
print(f"\n  ❌ Not shortlisted : {len(rejected)} employees")

# ── Ask to save to DB ─────────────────────────────────────────────────────────
print(f"\n{'='*65}")
save = input("  Save this shortlist to storage? (y/n): ").strip().lower()

if save == "y":
    prog_meta = {
        "name"          : prog_name,
        "min_qual_name" : QUAL_NAMES[min_qual],
        "min_exp"       : min_exp,
        "week_name"     : WEEK_NAMES[week_choice],
        "batch_size"    : batch_size,
        "waitlist_n"    : waitlist_n,
    }
    program_id = save_to_db(prog_meta, selected, waitlist)
    print(f"\n  ✅ Saved! Program ID : {program_id}")
    print(f"  📁 Stored in        : nominations.db")
    print(f"  📋 Records saved    : {len(selected)} selected + {len(waitlist)} waitlist")
else:
    print("\n  Shortlist not saved.")

# ── Ask to export to Excel ────────────────────────────────────────────────────
export = input("\n  Export shortlist to Excel? (y/n)      : ").strip().lower()

if export == "y":
    prog_meta = {
        "name"          : prog_name,
        "min_qual_name" : QUAL_NAMES[min_qual],
        "min_exp"       : min_exp,
        "week_name"     : WEEK_NAMES[week_choice],
        "batch_size"    : batch_size,
        "waitlist_n"    : waitlist_n,
    }
    file_path = export_to_excel(prog_name, prog_meta, selected, waitlist, rejected)
    print(f"\n  ✅ Excel exported!")
    print(f"  📂 Saved to : {file_path}")
    print(f"  📄 Sheets   : Summary · ✅ Selected · ⏳ Waitlist")
else:
    print("\n  Export skipped.")

print(f"\n{'='*65}")
print(f"  DONE — Awaiting manager approval")
print(f"{'='*65}")