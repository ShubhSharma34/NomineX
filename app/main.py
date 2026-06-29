from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import pandas as pd
import numpy as np
import joblib
import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse

app = FastAPI(title="Nomination Filtering System API")

# Allow React frontend to talk to this backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH  = os.path.join(BASE_DIR, "data",   "nomination_synthetic_dataset_v3.xlsx")
MODELS_DIR = os.path.join(BASE_DIR, "models")
DB_PATH    = os.path.join(BASE_DIR, "nominations.db")
EXPORTS    = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORTS, exist_ok=True)

# ── Load model once at startup ────────────────────────────────────────────────
model    = joblib.load(os.path.join(MODELS_DIR, "best_model.pkl"))
features = joblib.load(os.path.join(MODELS_DIR, "feature_names.pkl"))

# ── DB init ───────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS programs (
            program_id    INTEGER PRIMARY KEY AUTOINCREMENT,
            program_name  TEXT,
            min_qual      TEXT,
            min_exp       INTEGER,
            training_week TEXT,
            batch_size    INTEGER,
            waitlist_size INTEGER,
            run_at        TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS nominations (
            nomination_id INTEGER PRIMARY KEY AUTOINCREMENT,
            program_id    INTEGER,
            employee_id   TEXT,
            employee_name TEXT,
            department    TEXT,
            area_of_work  TEXT,
            qualification TEXT,
            experience    INTEGER,
            ai_score      REAL,
            fuzzy_match   REAL,
            status        TEXT,
            approval      TEXT DEFAULT 'Pending',
            rank          INTEGER,
            FOREIGN KEY (program_id) REFERENCES programs(program_id)
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ── Pydantic models ───────────────────────────────────────────────────────────
class ShortlistRequest(BaseModel):
    program_name  : str
    min_qual      : int
    min_exp       : int
    week_choice   : int
    batch_size    : int
    waitlist_n    : int = 5
    target_areas  : Optional[List[str]] = []

class ApprovalRequest(BaseModel):
    nomination_id : int
    decision      : str   # "Approved" or "Rejected"

class ColdStartRequest(BaseModel):
    name          : str
    experience    : int
    qual_input    : int
    skill_1_name  : str
    skill_1_level : int
    skill_2_name  : str
    skill_2_level : int
    skill_3_name  : Optional[str] = ""
    skill_3_level : Optional[int] = 0
    free_weeks    : List[int]
    perf_rating   : float
    months_since  : int = 24

# ── Helper: load & engineer features ─────────────────────────────────────────
def load_and_engineer():
    df_emp    = pd.read_excel(DATA_PATH, sheet_name="Employees")
    df_skills = pd.read_excel(DATA_PATH, sheet_name="Skills")
    df_hist   = pd.read_excel(DATA_PATH, sheet_name="Training History")
    df_avail  = pd.read_excel(DATA_PATH, sheet_name="Availability")

    QUAL_RANK = {"High School":1,"Diploma":2,"Bachelor's":3,"Master's":4,"MBA":5,"PhD":6}
    df_emp["qual_rank"]   = df_emp["Qualification"].map(QUAL_RANK).fillna(1)
    df_emp["experience"]  = df_emp["Experience (Years)"]
    df_emp["performance"] = df_emp["Performance Rating (1-5)"]
    df_emp["Last Promotion Date"] = pd.to_datetime(
        df_emp["Last Promotion Date"], dayfirst=True, errors="coerce")
    TODAY = pd.Timestamp("2025-06-05")
    df_emp["days_since_promotion"] = (TODAY - df_emp["Last Promotion Date"]).dt.days.fillna(365)
    dept_map = {d: i for i, d in enumerate(df_emp["Department"].unique())}
    df_emp["dept_code"] = df_emp["Department"].map(dept_map).fillna(0)

    hist_completed = df_hist[df_hist["Status"] == "Completed"]
    hist_failed    = df_hist[df_hist["Status"].isin(["Failed","Dropped"])]
    hist_agg = hist_completed.groupby("Employee ID").agg(
        trainings_completed=("Program Name","count"),
        avg_training_score=("Score (%)","mean")).reset_index()
    hist_fail_agg = hist_failed.groupby("Employee ID").agg(
        trainings_failed=("Program Name","count")).reset_index()

    LEVEL_RANK = {"Beginner":1,"Intermediate":2,"Advanced":3}
    for col in ["Skill 1 Level","Skill 2 Level","Skill 3 Level"]:
        df_skills[col] = df_skills[col].map(LEVEL_RANK).fillna(1)
    df_skills["avg_skill_level"] = df_skills[["Skill 1 Level","Skill 2 Level","Skill 3 Level"]].mean(axis=1)
    df_skills["max_skill_level"] = df_skills[["Skill 1 Level","Skill 2 Level","Skill 3 Level"]].max(axis=1)
    skill_agg = df_skills[["Employee ID","avg_skill_level","max_skill_level"]]

    avail_cols = [c for c in df_avail.columns if c != "Employee ID"]
    for col in avail_cols:
        df_avail[col] = df_avail[col].map({"Yes":1,"No":0}).fillna(0)
    df_avail["weeks_available"] = df_avail[avail_cols].sum(axis=1)

    df = df_emp[["Employee ID","Name","Department","Area of Work","Qualification",
                 "experience","qual_rank","performance","days_since_promotion","dept_code"]].copy()
    df = df.merge(hist_agg,      on="Employee ID", how="left")
    df = df.merge(hist_fail_agg, on="Employee ID", how="left")
    df = df.merge(skill_agg,     on="Employee ID", how="left")
    df = df.merge(df_avail[["Employee ID","weeks_available"] + avail_cols],
                  on="Employee ID", how="left")
    df = df.fillna(0)
    return df, avail_cols

# ── Fuzzy functions ───────────────────────────────────────────────────────────
def fuzzy_qual(q, mq):
    gap = q - mq
    return 100.0 if gap>=0 else (60.0 if gap==-1 else (30.0 if gap==-2 else 10.0))

def fuzzy_exp(e, me):
    if e >= me or me == 0: return 100.0
    return round(max(10.0, (e/me)**1.5*100), 1)

def fuzzy_avail(free, weeks):
    return 100.0 if free==1 else (30.0 if weeks>0 else 0.0)

def fuzzy_skill(avg, mx):
    return round((mx/3)*60 + (avg/3)*40, 1)

def fuzzy_train(tc, ts, tf):
    return round(max(0.0, min(tc/4,1.0)*50 + (ts/100)*35 - min(tf*10,30)), 1)

def fuzzy_area(area, target_areas):
    if not target_areas: return 50.0
    return 100.0 if str(area).lower() in [a.lower() for a in target_areas] else 20.0

# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"message": "Nomination Filtering System API is running"}

# ── Stats for dashboard ───────────────────────────────────────────────────────
@app.get("/stats")
def get_stats():
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("SELECT COUNT(*) FROM programs")
    total_programs = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM nominations WHERE status='Selected'")
    total_selected = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM nominations WHERE approval='Pending' AND status='Selected'")
    pending = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM nominations WHERE approval='Approved'")
    approved = c.fetchone()[0]
    conn.close()
    return {
        "total_employees" : 2000,
        "total_programs"  : total_programs,
        "total_selected"  : total_selected,
        "pending_approval": pending,
        "approved"        : approved,
    }

# ── Get all programs (history) ────────────────────────────────────────────────
@app.get("/programs")
def get_programs():
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("SELECT * FROM programs ORDER BY run_at DESC")
    rows = c.fetchall()
    conn.close()
    return [{"program_id": r[0], "program_name": r[1], "min_qual": r[2],
             "min_exp": r[3], "training_week": r[4], "batch_size": r[5],
             "waitlist_size": r[6], "run_at": r[7]} for r in rows]

# ── Get nominations for a program ─────────────────────────────────────────────
@app.get("/nominations/{program_id}")
def get_nominations(program_id: int):
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("SELECT * FROM nominations WHERE program_id=? ORDER BY status, rank",
              (program_id,))
    rows = c.fetchall()
    conn.close()
    return [{"nomination_id": r[0], "program_id": r[1], "employee_id": r[2],
             "employee_name": r[3], "department": r[4], "area_of_work": r[5],
             "qualification": r[6], "experience": r[7], "ai_score": r[8],
             "fuzzy_match": r[9], "status": r[10], "approval": r[11],
             "rank": r[12]} for r in rows]

# ── Generate shortlist ────────────────────────────────────────────────────────
@app.post("/shortlist")
def generate_shortlist(req: ShortlistRequest):
    QUAL_NAMES = {1:"High School",2:"Diploma",3:"Bachelor's",4:"Master's",5:"MBA",6:"PhD"}
    WEEK_NAMES = {1:"Week 1 (Jun 9-13)",2:"Week 2 (Jun 16-20)",
                  3:"Week 3 (Jun 23-27)",4:"Week 4 (Jun 30-Jul 4)"}

    df, avail_cols = load_and_engineer()
    week_col = avail_cols[req.week_choice - 1]

    # Area filter
    if req.target_areas:
        df = df[df["Area of Work"].str.lower().isin(
            [a.lower() for a in req.target_areas])].copy()

    # Hard filter
    df_eligible = df[
        (df[week_col]      == 1) &
        (df["qual_rank"]   >= req.min_qual) &
        (df["experience"]  >= req.min_exp)
    ].copy()

    if len(df_eligible) == 0:
        raise HTTPException(status_code=404, detail="No eligible employees found")

    batch = min(req.batch_size, len(df_eligible))
    wait  = min(req.waitlist_n, max(0, len(df_eligible) - batch))

    # AI score
    df_eligible["ai_score"] = np.clip(model.predict(df_eligible[features]), 10, 95).round(2)

    # Fuzzy match
    df_eligible["fz_qual"]  = df_eligible.apply(lambda r: fuzzy_qual(r["qual_rank"], req.min_qual), axis=1)
    df_eligible["fz_exp"]   = df_eligible.apply(lambda r: fuzzy_exp(r["experience"], req.min_exp), axis=1)
    df_eligible["fz_avail"] = df_eligible.apply(lambda r: fuzzy_avail(r[week_col], r["weeks_available"]), axis=1)
    df_eligible["fz_skill"] = df_eligible.apply(lambda r: fuzzy_skill(r["avg_skill_level"], r["max_skill_level"]), axis=1)
    df_eligible["fz_train"] = df_eligible.apply(lambda r: fuzzy_train(r["trainings_completed"], r["avg_training_score"], r["trainings_failed"]), axis=1)
    df_eligible["fz_area"]  = df_eligible.apply(lambda r: fuzzy_area(r["Area of Work"], req.target_areas), axis=1)
    df_eligible["fuzzy_match"] = (
        df_eligible["fz_qual"]  * 0.18 +
        df_eligible["fz_exp"]   * 0.22 +
        df_eligible["fz_avail"] * 0.18 +
        df_eligible["fz_skill"] * 0.18 +
        df_eligible["fz_train"] * 0.12 +
        df_eligible["fz_area"]  * 0.12
    ).round(1)

    df_eligible = df_eligible.sort_values("ai_score", ascending=False).reset_index(drop=True)

    def make_list(subset, status):
        result = []
        for rank, (_, r) in enumerate(subset.iterrows(), start=1):
            result.append({
                "employee_id"  : r["Employee ID"],
                "employee_name": r["Name"],
                "department"   : r["Department"],
                "area_of_work" : r["Area of Work"],
                "qualification": r["Qualification"],
                "experience"   : int(r["experience"]),
                "ai_score"     : float(r["ai_score"]),
                "fuzzy_match"  : float(r["fuzzy_match"]),
                "fz_qual"      : float(r["fz_qual"]),
                "fz_exp"       : float(r["fz_exp"]),
                "fz_avail"     : float(r["fz_avail"]),
                "fz_skill"     : float(r["fz_skill"]),
                "fz_train"     : float(r["fz_train"]),
                "fz_area"      : float(r["fz_area"]),
                "status"       : status,
                "rank"         : rank,
            })
        return result

    selected = df_eligible.iloc[:batch]
    waitlist = df_eligible.iloc[batch:batch+wait]
    rejected = df_eligible.iloc[batch+wait:]

    return {
        "program_name"   : req.program_name,
        "training_week"  : WEEK_NAMES[req.week_choice],
        "total_eligible" : len(df_eligible),
        "selected"       : make_list(selected, "Selected"),
        "waitlist"       : make_list(waitlist, "Waitlist"),
        "rejected_count" : len(rejected),
    }

# ── Save shortlist to DB ──────────────────────────────────────────────────────
@app.post("/save")
def save_shortlist(data: dict):
    QUAL_NAMES = {1:"High School",2:"Diploma",3:"Bachelor's",4:"Master's",5:"MBA",6:"PhD"}
    WEEK_NAMES = {1:"Week 1 (Jun 9-13)",2:"Week 2 (Jun 16-20)",
                  3:"Week 3 (Jun 23-27)",4:"Week 4 (Jun 30-Jul 4)"}
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("""
        INSERT INTO programs
        (program_name, min_qual, min_exp, training_week, batch_size, waitlist_size, run_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (data["program_name"], QUAL_NAMES.get(data["min_qual"], ""),
          data["min_exp"], WEEK_NAMES.get(data["week_choice"], ""),
          data["batch_size"], data["waitlist_n"],
          datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    program_id = c.lastrowid
    for emp in data.get("selected", []) + data.get("waitlist", []):
        c.execute("""
            INSERT INTO nominations
            (program_id, employee_id, employee_name, department, area_of_work,
             qualification, experience, ai_score, fuzzy_match, status, rank)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (program_id, emp["employee_id"], emp["employee_name"],
              emp["department"], emp["area_of_work"], emp["qualification"],
              emp["experience"], emp["ai_score"], emp["fuzzy_match"],
              emp["status"], emp["rank"]))
    conn.commit()
    conn.close()
    return {"program_id": program_id, "message": "Saved successfully"}

# ── Approve / Reject nomination ───────────────────────────────────────────────
@app.post("/approve")
def approve_nomination(req: ApprovalRequest):
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    c.execute("UPDATE nominations SET approval=? WHERE nomination_id=?",
              (req.decision, req.nomination_id))
    # If rejected, promote top waitlist person
    if req.decision == "Rejected":
        c.execute("""
            SELECT program_id FROM nominations WHERE nomination_id=?
        """, (req.nomination_id,))
        row = c.fetchone()
        if row:
            program_id = row[0]
            c.execute("""
                SELECT nomination_id FROM nominations
                WHERE program_id=? AND status='Waitlist' AND approval='Pending'
                ORDER BY rank LIMIT 1
            """, (program_id,))
            promote = c.fetchone()
            if promote:
                c.execute("""
                    UPDATE nominations SET status='Selected', approval='Promoted'
                    WHERE nomination_id=?
                """, (promote[0],))
    conn.commit()
    conn.close()
    return {"message": f"Nomination {req.decision}"}

# ── Cold Start ────────────────────────────────────────────────────────────────
@app.post("/coldstart")
def cold_start(req: ColdStartRequest):
    df_prog = pd.read_excel(DATA_PATH, sheet_name="Training Programs")

    all_levels = [req.skill_1_level, req.skill_2_level]
    if req.skill_3_level and req.skill_3_level > 0:
        all_levels.append(req.skill_3_level)
    avg_skill = round(np.mean(all_levels), 2)
    max_skill = max(all_levels)

    emp_vector = {
        "experience"          : req.experience,
        "qual_rank"           : req.qual_input,
        "performance"         : req.perf_rating,
        "days_since_promotion": req.months_since * 30,
        "dept_code"           : 4,
        "trainings_completed" : max(0, 3 - req.months_since // 6),
        "avg_training_score"  : 70.0,
        "trainings_failed"    : 0,
        "avg_skill_level"     : avg_skill,
        "max_skill_level"     : max_skill,
        "weeks_available"     : len(req.free_weeks),
    }
    X = pd.DataFrame([emp_vector])[features]
    ai_score = float(np.clip(model.predict(X)[0], 10, 95))

    emp_skills = {req.skill_1_name.lower(): req.skill_1_level,
                  req.skill_2_name.lower(): req.skill_2_level}
    if req.skill_3_name and req.skill_3_level:
        emp_skills[req.skill_3_name.lower()] = req.skill_3_level

    QUAL_THRESH = {"High School":1,"Diploma":2,"Bachelor's":3,"Master's":4,"MBA":5,"PhD":6}
    WEEK_MAP    = {1:0, 2:1, 3:2, 4:3}

    results = []
    for _, prog in df_prog.iterrows():
        prog_min_qual = QUAL_THRESH.get(prog["Min Qualification"], 1)
        prog_min_exp  = prog["Min Experience (Yrs)"]
        prog_week_num = int(str(prog["Training Week"]).split()[-1]) if "Week" in str(prog["Training Week"]) else 1
        prog_skill    = str(prog["Skill Required"]).lower()

        qual_gap   = req.qual_input - prog_min_qual
        qual_score = 100.0 if qual_gap>=0 else (60.0 if qual_gap==-1 else (30.0 if qual_gap==-2 else 0.0))
        exp_score  = 100.0 if req.experience >= prog_min_exp else round(max(0.0,(req.experience/max(prog_min_exp,1))**1.5*100),1)
        avail_score= 100.0 if prog_week_num in req.free_weeks else 0.0
        matched    = emp_skills.get(prog_skill, 0)
        skill_score= {3:100.0, 2:70.0, 1:40.0}.get(matched, 20.0)
        growth     = {0:100.0, 1:80.0, 2:50.0, 3:20.0}.get(matched, 50.0)

        compat = (qual_score*0.20 + exp_score*0.20 + avail_score*0.25 +
                  skill_score*0.20 + growth*0.15)
        if avail_score == 0:
            compat = min(compat, 30.0)

        results.append({
            "program_id"    : str(prog["Program ID"]),
            "program_name"  : prog["Program Name"],
            "skill_required": prog["Skill Required"],
            "min_qual"      : prog["Min Qualification"],
            "min_exp"       : int(prog["Min Experience (Yrs)"]),
            "week"          : prog["Training Week"],
            "batch_size"    : int(prog["Batch Size"]),
            "qual_score"    : qual_score,
            "exp_score"     : exp_score,
            "avail_score"   : avail_score,
            "skill_score"   : skill_score,
            "growth_score"  : growth,
            "compatibility" : round(compat, 1),
            "eligible"      : qual_score > 0 and exp_score >= 50,
            "available"     : avail_score == 100.0,
        })

    results.sort(key=lambda x: x["compatibility"], reverse=True)
    return {"employee_name": req.name, "ai_score": round(ai_score, 1), "programs": results}
@app.post("/export")
def export_to_excel(data: dict):
    from fastapi.responses import FileResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prog_name = data.get("program_name", "Shortlist")
    selected  = data.get("selected", [])
    waitlist  = data.get("waitlist", [])

    def hdr(cell, bg="1F4E79"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.merge_cells("A1:B1")
    ws["A1"] = "NOMINATION FILTERING SYSTEM"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:B2")
    ws["A2"] = f"Shortlist Report — {prog_name}"
    ws["A2"].font = Font(bold=True, size=11, color="2E75B6", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    details = [
        ("Program Name",             prog_name),
        ("Generated On",             datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("Total Selected",           len(selected)),
        ("Total Waitlisted",         len(waitlist)),
        ("Avg AI Score (Selected)",  f"{sum(e['ai_score'] for e in selected)/len(selected):.2f}" if selected else "N/A"),
        ("Avg Fuzzy Match (Selected)", f"{sum(e['fuzzy_match'] for e in selected)/len(selected):.1f}%" if selected else "N/A"),
    ]
    for label, value in details:
        ws.append([label, value])
        row = ws.max_row
        ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="404040")
        ws[f"B{row}"].font = Font(name="Arial", size=10)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="F2F2F2")
        ws.row_dimensions[row].height = 16

    # ── Selected sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("✅ Selected")
    headers = ["Rank","Employee ID","Name","Department","Area of Work",
               "Qualification","Exp (Yrs)","AI Score",
               "Qual%","Exp%","Avail%","Skill%","Train%","Area%","Match%","Fit"]
    widths  = [6,13,22,14,22,14,9,11,8,8,8,8,8,8,10,14]

    ws2.append(headers)
    for i, w in enumerate(widths, 1):
        hdr(ws2.cell(row=1, column=i))
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    def fit_text(pct):
        if pct >= 75:   return "Strong Fit"
        elif pct >= 50: return "Moderate Fit"
        elif pct >= 30: return "Weak Fit"
        else:           return "Poor Fit"

    FIT_COLORS = {
        "Strong Fit":   ("14532D","DCFCE7"),
        "Moderate Fit": ("713F12","FEF9C3"),
        "Weak Fit":     ("7C2D12","FFEDD5"),
        "Poor Fit":     ("7F1D1D","FEE2E2"),
    }

    for rank, emp in enumerate(selected, 1):
        fit = fit_text(emp.get("fuzzy_match", 0))
        row_data = [
            rank, emp.get("employee_id",""), emp.get("employee_name",""),
            emp.get("department",""), emp.get("area_of_work",""),
            emp.get("qualification",""), emp.get("experience",0),
            round(emp.get("ai_score",0), 2),
            f"{emp.get('fz_qual',0):.0f}%", f"{emp.get('fz_exp',0):.0f}%",
            f"{emp.get('fz_avail',0):.0f}%", f"{emp.get('fz_skill',0):.0f}%",
            f"{emp.get('fz_train',0):.0f}%", f"{emp.get('fz_area',0):.0f}%",
            f"{emp.get('fuzzy_match',0):.1f}%", fit
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row
        bg = "FFFFFF" if rank % 2 == 0 else "F7FAFF"
        for col_idx in range(1, len(row_data)+1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left",
                                       vertical="center")
            cell.border    = thin()
        fit_cell = ws2.cell(row=row_idx, column=16)
        fc, bc   = FIT_COLORS.get(fit, ("000000","FFFFFF"))
        fit_cell.fill = PatternFill("solid", fgColor=bc)
        fit_cell.font = Font(name="Arial", size=10, color=fc, bold=True)
        ws2.cell(row=row_idx, column=8).font = Font(name="Arial", size=10,
                                                     bold=True, color="1F4E79")
        ws2.row_dimensions[row_idx].height = 16

    # ── Waitlist sheet ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("⏳ Waitlist")
    ws3.append(headers)
    for i, w in enumerate(widths, 1):
        hdr(ws3.cell(row=1, column=i), bg="7B3F00")
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    for rank, emp in enumerate(waitlist, 1):
        fit = fit_text(emp.get("fuzzy_match", 0))
        row_data = [
            rank, emp.get("employee_id",""), emp.get("employee_name",""),
            emp.get("department",""), emp.get("area_of_work",""),
            emp.get("qualification",""), emp.get("experience",0),
            round(emp.get("ai_score",0), 2),
            f"{emp.get('fz_qual',0):.0f}%", f"{emp.get('fz_exp',0):.0f}%",
            f"{emp.get('fz_avail',0):.0f}%", f"{emp.get('fz_skill',0):.0f}%",
            f"{emp.get('fz_train',0):.0f}%", f"{emp.get('fz_area',0):.0f}%",
            f"{emp.get('fuzzy_match',0):.1f}%", fit
        ]
        ws3.append(row_data)
        row_idx = ws3.max_row
        bg = "FFFFFF" if rank % 2 == 0 else "FFFBF5"
        for col_idx in range(1, len(row_data)+1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left",
                                       vertical="center")
            cell.border    = thin()
        ws3.row_dimensions[row_idx].height = 16

    # Save file
    os.makedirs(EXPORTS, exist_ok=True)
    safe   = "".join(c if c.isalnum() or c in " _-" else "_" for c in prog_name)
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    fpath  = os.path.join(EXPORTS, f"Shortlist_{safe}_{ts}.xlsx")
    wb.save(fpath)
    return FileResponse(
        fpath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"Shortlist_{safe}_{ts}.xlsx"
    )